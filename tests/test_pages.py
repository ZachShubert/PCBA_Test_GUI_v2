"""
Test Suite for PCBA Database Application Pages

Tests the main functionality of:
- Database Page: CRUD operations, filtering, navigation
- Reports Page: Data generation, export configuration
- Search Page: Search, autocomplete, compare mode

Run with: pytest tests/test_pages.py -v
Or without pytest: python tests/test_pages.py

Note: Requires SQLAlchemy and openpyxl to be installed.
      Install with: pip install -r requirements.txt
"""
import sys
import os
import unittest
import tempfile
import shutil
from datetime import datetime, timedelta
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Check for required dependencies
SQLALCHEMY_AVAILABLE = False
OPENPYXL_AVAILABLE = False

try:
    import sqlalchemy
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    print("WARNING: SQLAlchemy not installed. Database tests will be skipped.")
    print("         Install with: pip install sqlalchemy")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("WARNING: openpyxl not installed. Excel export tests will be skipped.")
    print("         Install with: pip install openpyxl")

# Only import database modules if SQLAlchemy is available
if SQLALCHEMY_AVAILABLE:
    from src.database import DatabaseManager
    from src.database.database_device_tables import PCBABoard, PMT
    from src.database.database_test_log_tables import TestLog, SubTest, Spec, MeasurementType
    from src.database.database_manufacturer_tables import Manufacturer, ManufacturerSpec


@unittest.skipUnless(SQLALCHEMY_AVAILABLE, "SQLAlchemy not installed")
class TestDatabaseSetup(unittest.TestCase):
    """Test database initialization and basic operations."""
    
    @classmethod
    def setUpClass(cls):
        """Create a temporary database for testing."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.temp_dir, "test_database.db")
        cls.db_url = f"sqlite:///{cls.db_path}"
        cls.db = DatabaseManager(cls.db_url)
    
    @classmethod
    def tearDownClass(cls):
        """Clean up temporary database."""
        cls.db.close()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_database_initialization(self):
        """Test that database initializes correctly."""
        self.assertIsNotNone(self.db)
        self.assertTrue(os.path.exists(self.db_path))
        print("✓ Database initialization passed")
    
    def test_02_add_pia_board(self):
        """Test adding a PIA board."""
        with self.db.session_scope() as session:
            board = PCBABoard(
                serial_number="TEST-PIA-001",
                part_number="PART-A",
                generation_project="Gen1",
                version="1.0"
            )
            session.add(board)
        
        # Verify
        with self.db.session_scope() as session:
            board = session.query(PCBABoard).filter_by(serial_number="TEST-PIA-001").first()
            self.assertIsNotNone(board)
            self.assertEqual(board.part_number, "PART-A")
        print("✓ Add PIA board passed")
    
    def test_03_add_pmt_device(self):
        """Test adding a PMT device."""
        with self.db.session_scope() as session:
            pmt = PMT(
                pmt_serial_number="TEST-PMT-001",
                generation="Gen1",
                batch_number="BATCH-A"
            )
            session.add(pmt)
        
        # Verify
        with self.db.session_scope() as session:
            pmt = session.query(PMT).filter_by(pmt_serial_number="TEST-PMT-001").first()
            self.assertIsNotNone(pmt)
            self.assertEqual(pmt.batch_number, "BATCH-A")
        print("✓ Add PMT device passed")
    
    def test_04_add_test_log_with_specs(self):
        """Test adding a test log with subtests and specs."""
        with self.db.session_scope() as session:
            # Get the board and PMT
            board = session.query(PCBABoard).filter_by(serial_number="TEST-PIA-001").first()
            pmt = session.query(PMT).filter_by(pmt_serial_number="TEST-PMT-001").first()
            
            # Create test log
            test_log = TestLog(
                pia_board_id=board.id,
                pmt_id=pmt.id,
                name="Voltage Test",
                description="Testing voltage outputs",
                test_fixture="Test Fixture A",
                full_test_completed=True,
                full_test_passed=True,
                html_content="<html><body><h1>Test Report</h1><p>All tests passed.</p></body></html>"
            )
            session.add(test_log)
            session.flush()
            
            # Create subtest
            subtest = SubTest(
                test_log_id=test_log.id,
                name="Power Supply Test",
                description="Testing power supply voltages"
            )
            session.add(subtest)
            session.flush()
            
            # Create specs
            specs = [
                Spec(
                    sub_test_id=subtest.id,
                    name="Output Voltage 5V",
                    unit="V",
                    measurement_type=MeasurementType.FLOAT,
                    measurement=5.02,
                    lower_limit=4.75,
                    upper_limit=5.25,
                    result=True
                ),
                Spec(
                    sub_test_id=subtest.id,
                    name="Output Voltage 3.3V",
                    unit="V",
                    measurement_type=MeasurementType.FLOAT,
                    measurement=3.28,
                    lower_limit=3.13,
                    upper_limit=3.47,
                    result=True
                ),
                Spec(
                    sub_test_id=subtest.id,
                    name="Output Voltage 12V",
                    unit="V",
                    measurement_type=MeasurementType.FLOAT,
                    measurement=12.15,
                    lower_limit=11.4,
                    upper_limit=12.6,
                    result=True
                )
            ]
            for spec in specs:
                session.add(spec)
        
        # Verify
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).filter_by(name="Voltage Test").first()
            self.assertIsNotNone(test_log)
            self.assertTrue(test_log.full_test_passed)
            self.assertEqual(len(test_log.sub_tests), 1)
            self.assertEqual(len(test_log.sub_tests[0].specs), 3)
        print("✓ Add test log with specs passed")
    
    def test_05_add_manufacturer_data(self):
        """Test adding manufacturer data."""
        with self.db.session_scope() as session:
            mfr = Manufacturer(
                name="Test Manufacturer",
                description="A test manufacturer",
                website="https://example.com"
            )
            session.add(mfr)
            session.flush()
            
            # Add spec
            spec = ManufacturerSpec(
                manufacturer_id=mfr.id,
                spec_name="Output Voltage 5V",
                measurement=5.00,
                unit="V",
                lower_limit=4.75,
                upper_limit=5.25
            )
            session.add(spec)
        
        # Verify
        with self.db.session_scope() as session:
            mfr = session.query(Manufacturer).filter_by(name="Test Manufacturer").first()
            self.assertIsNotNone(mfr)
            self.assertEqual(len(mfr.specs), 1)
        print("✓ Add manufacturer data passed")


@unittest.skipUnless(SQLALCHEMY_AVAILABLE, "SQLAlchemy not installed")
class TestDatabaseQueries(unittest.TestCase):
    """Test database query functionality."""
    
    @classmethod
    def setUpClass(cls):
        """Create a temporary database with test data."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.temp_dir, "test_queries.db")
        cls.db_url = f"sqlite:///{cls.db_path}"
        cls.db = DatabaseManager(cls.db_url)
        
        # Populate test data
        cls._populate_test_data()
    
    @classmethod
    def _populate_test_data(cls):
        """Populate database with test data."""
        with cls.db.session_scope() as session:
            # Create multiple boards
            boards = []
            for i in range(5):
                board = PCBABoard(
                    serial_number=f"SN-{i:04d}",
                    part_number=f"PART-{'A' if i < 3 else 'B'}",
                    generation_project="Gen1"
                )
                session.add(board)
                boards.append(board)
            session.flush()
            
            # Create PMTs
            pmts = []
            for i in range(3):
                pmt = PMT(
                    pmt_serial_number=f"PMT-{i:04d}",
                    batch_number=f"BATCH-{'X' if i < 2 else 'Y'}",
                    generation="Gen1"
                )
                session.add(pmt)
                pmts.append(pmt)
            session.flush()
            
            # Create test logs with varying results
            fixtures = ["Fixture A", "Fixture B", "Fixture A"]
            for i, board in enumerate(boards[:3]):
                for j in range(2):  # 2 tests per board
                    test_log = TestLog(
                        pia_board_id=board.id,
                        pmt_id=pmts[i % len(pmts)].id,
                        name=f"Test Run {i}-{j}",
                        test_fixture=fixtures[i % len(fixtures)],
                        full_test_completed=True,
                        full_test_passed=(j == 0),  # First test passes, second fails
                        created_at=datetime.now() - timedelta(days=i*10 + j),
                        html_content=f"<html><body><h1>Test {i}-{j}</h1></body></html>"
                    )
                    session.add(test_log)
                    session.flush()
                    
                    subtest = SubTest(
                        test_log_id=test_log.id,
                        name="Voltage Test"
                    )
                    session.add(subtest)
                    session.flush()
                    
                    spec = Spec(
                        sub_test_id=subtest.id,
                        name="Output Voltage 5V",
                        unit="V",
                        measurement_type=MeasurementType.FLOAT,
                        measurement=5.0 + (0.1 * i) + (0.05 * j),
                        lower_limit=4.75,
                        upper_limit=5.25,
                        result=(j == 0)
                    )
                    session.add(spec)
    
    @classmethod
    def tearDownClass(cls):
        """Clean up temporary database."""
        cls.db.close()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_get_all_serial_numbers(self):
        """Test retrieving all PIA serial numbers."""
        serials = self.db.queries.pias.get_all_serial_numbers()
        self.assertEqual(len(serials), 5)
        self.assertIn("SN-0000", serials)
        print("✓ Get all serial numbers passed")
    
    def test_02_get_all_spec_names(self):
        """Test retrieving all spec names."""
        spec_names = self.db.queries.specs.get_all_spec_names()
        self.assertIn("Output Voltage 5V", spec_names)
        print("✓ Get all spec names passed")
    
    def test_03_search_test_logs(self):
        """Test searching test logs."""
        results = self.db.search("SN-0000")
        self.assertGreater(len(results), 0)
        print("✓ Search test logs passed")
    
    def test_04_get_database_stats(self):
        """Test getting database statistics."""
        stats = self.db.get_database_stats()
        self.assertEqual(stats['total_boards'], 5)
        self.assertEqual(stats['total_pmts'], 3)
        self.assertGreater(stats['total_test_logs'], 0)
        print("✓ Get database stats passed")
    
    def test_05_filter_by_fixture(self):
        """Test filtering specs by test fixture."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            
            # Count test logs for Fixture A
            count = session.query(TestLog).filter(
                TestLog.test_fixture == "Fixture A"
            ).count()
            
            self.assertGreater(count, 0)
        print("✓ Filter by fixture passed")


@unittest.skipUnless(SQLALCHEMY_AVAILABLE, "SQLAlchemy not installed")
class TestDatabasePageLogic(unittest.TestCase):
    """Test Database Page logic without GUI."""
    
    @classmethod
    def setUpClass(cls):
        """Create a temporary database with test data."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.temp_dir, "test_db_page.db")
        cls.db_url = f"sqlite:///{cls.db_path}"
        cls.db = DatabaseManager(cls.db_url)
        
        # Add test data
        with cls.db.session_scope() as session:
            board = PCBABoard(serial_number="DB-TEST-001", part_number="PART-X")
            session.add(board)
            session.flush()
            
            test_log = TestLog(
                pia_board_id=board.id,
                name="DB Page Test",
                test_fixture="Fixture A",
                full_test_completed=True,
                full_test_passed=True,
                html_content="<html><body>Test</body></html>"
            )
            session.add(test_log)
    
    @classmethod
    def tearDownClass(cls):
        cls.db.close()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_load_test_logs(self):
        """Test loading test logs with filters."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            from sqlalchemy import desc
            
            query = session.query(TestLog).options(
                joinedload(TestLog.pia_board),
                joinedload(TestLog.pmt_device)
            ).order_by(desc(TestLog.created_at))
            
            results = query.all()
            self.assertGreater(len(results), 0)
            
            # Check that related data is loaded
            for tl in results:
                self.assertIsNotNone(tl.pia_board)
        print("✓ Load test logs passed")
    
    def test_02_update_test_log_metadata(self):
        """Test updating test log metadata."""
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).filter_by(name="DB Page Test").first()
            self.assertIsNotNone(test_log)
            
            # Update fixture
            test_log.test_fixture = "Fixture B"
        
        # Verify
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).filter_by(name="DB Page Test").first()
            self.assertEqual(test_log.test_fixture, "Fixture B")
        print("✓ Update test log metadata passed")
    
    def test_03_update_device_info(self):
        """Test updating device serial/part numbers."""
        with self.db.session_scope() as session:
            board = session.query(PCBABoard).filter_by(serial_number="DB-TEST-001").first()
            self.assertIsNotNone(board)
            
            # Update part number
            board.part_number = "PART-Y"
        
        # Verify
        with self.db.session_scope() as session:
            board = session.query(PCBABoard).filter_by(serial_number="DB-TEST-001").first()
            self.assertEqual(board.part_number, "PART-Y")
        print("✓ Update device info passed")
    
    def test_04_delete_test_log(self):
        """Test deleting a test log."""
        # Add a test log to delete
        with self.db.session_scope() as session:
            board = session.query(PCBABoard).first()
            test_log = TestLog(
                pia_board_id=board.id,
                name="To Delete",
                test_fixture="Fixture X"
            )
            session.add(test_log)
        
        # Delete it
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).filter_by(name="To Delete").first()
            self.assertIsNotNone(test_log)
            session.delete(test_log)
        
        # Verify deletion
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).filter_by(name="To Delete").first()
            self.assertIsNone(test_log)
        print("✓ Delete test log passed")


@unittest.skipUnless(SQLALCHEMY_AVAILABLE, "SQLAlchemy not installed")
class TestReportsPageLogic(unittest.TestCase):
    """Test Reports Page logic without GUI."""
    
    @classmethod
    def setUpClass(cls):
        """Create a temporary database with test data."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.temp_dir, "test_reports.db")
        cls.db_url = f"sqlite:///{cls.db_path}"
        cls.db = DatabaseManager(cls.db_url)
        
        # Add comprehensive test data
        with cls.db.session_scope() as session:
            # Create boards
            for i in range(3):
                board = PCBABoard(
                    serial_number=f"RPT-SN-{i:03d}",
                    part_number="RPT-PART"
                )
                session.add(board)
            session.flush()
            
            boards = session.query(PCBABoard).all()
            
            # Create test logs with specs
            for board in boards:
                for test_num in range(2):
                    test_log = TestLog(
                        pia_board_id=board.id,
                        name=f"Report Test {test_num}",
                        test_fixture="Fixture A",
                        full_test_completed=True,
                        full_test_passed=True,
                        created_at=datetime.now() - timedelta(days=test_num * 7)
                    )
                    session.add(test_log)
                    session.flush()
                    
                    subtest = SubTest(test_log_id=test_log.id, name="Voltage Test")
                    session.add(subtest)
                    session.flush()
                    
                    # Add multiple specs
                    for voltage in [5, 3.3, 12]:
                        spec = Spec(
                            sub_test_id=subtest.id,
                            name=f"Output Voltage {voltage}V",
                            unit="V",
                            measurement_type=MeasurementType.FLOAT,
                            measurement=voltage + 0.02,
                            lower_limit=voltage * 0.95,
                            upper_limit=voltage * 1.05,
                            result=True
                        )
                        session.add(spec)
    
    @classmethod
    def tearDownClass(cls):
        cls.db.close()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_query_specs_by_name(self):
        """Test querying specs by name."""
        with self.db.session_scope() as session:
            specs = session.query(Spec).filter(Spec.name == "Output Voltage 5V").all()
            self.assertGreater(len(specs), 0)
            
            for spec in specs:
                self.assertAlmostEqual(spec.measurement, 5.02, places=1)
        print("✓ Query specs by name passed")
    
    def test_02_query_specs_with_filters(self):
        """Test querying specs with device filters."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            
            specs = session.query(Spec).options(
                joinedload(Spec.sub_test).joinedload(SubTest.test_log).joinedload(TestLog.pia_board)
            ).join(Spec.sub_test).join(SubTest.test_log).join(TestLog.pia_board).filter(
                PCBABoard.serial_number == "RPT-SN-000"
            ).all()
            
            self.assertGreater(len(specs), 0)
            for spec in specs:
                self.assertEqual(
                    spec.sub_test.test_log.pia_board.serial_number,
                    "RPT-SN-000"
                )
        print("✓ Query specs with filters passed")
    
    def test_03_group_specs_by_device(self):
        """Test grouping specs by device."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            from collections import defaultdict
            
            specs = session.query(Spec).options(
                joinedload(Spec.sub_test).joinedload(SubTest.test_log).joinedload(TestLog.pia_board)
            ).filter(Spec.name == "Output Voltage 5V").all()
            
            # Group by device
            device_groups = defaultdict(list)
            for spec in specs:
                serial = spec.sub_test.test_log.pia_board.serial_number
                device_groups[serial].append(spec)
            
            self.assertEqual(len(device_groups), 3)  # 3 boards
            for serial, group in device_groups.items():
                self.assertEqual(len(group), 2)  # 2 tests per board
        print("✓ Group specs by device passed")
    
    def test_04_calculate_pass_fail(self):
        """Test pass/fail calculation based on limits."""
        with self.db.session_scope() as session:
            specs = session.query(Spec).filter(Spec.name == "Output Voltage 5V").all()
            
            for spec in specs:
                # Check if measurement is within limits
                if spec.lower_limit is not None and spec.upper_limit is not None:
                    in_limits = spec.lower_limit <= spec.measurement <= spec.upper_limit
                    self.assertTrue(in_limits)
        print("✓ Calculate pass/fail passed")


@unittest.skipUnless(SQLALCHEMY_AVAILABLE, "SQLAlchemy not installed")
class TestSearchPageLogic(unittest.TestCase):
    """Test Search Page logic without GUI."""
    
    @classmethod
    def setUpClass(cls):
        """Create a temporary database with test data."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.db_path = os.path.join(cls.temp_dir, "test_search.db")
        cls.db_url = f"sqlite:///{cls.db_path}"
        cls.db = DatabaseManager(cls.db_url)
        
        # Add test data
        with cls.db.session_scope() as session:
            # Create boards with various serials/parts
            serials = ["ABC-001", "ABC-002", "XYZ-001", "XYZ-002"]
            parts = ["PART-A", "PART-A", "PART-B", "PART-B"]
            
            for serial, part in zip(serials, parts):
                board = PCBABoard(serial_number=serial, part_number=part)
                session.add(board)
            session.flush()
            
            # Create PMTs
            pmt_serials = ["PMT-111", "PMT-222"]
            for pmt_serial in pmt_serials:
                pmt = PMT(pmt_serial_number=pmt_serial, batch_number="BATCH-1")
                session.add(pmt)
            session.flush()
            
            # Create test logs
            boards = session.query(PCBABoard).all()
            pmts = session.query(PMT).all()
            
            for i, board in enumerate(boards):
                test_log = TestLog(
                    pia_board_id=board.id,
                    pmt_id=pmts[i % len(pmts)].id,
                    name=f"Search Test {i}",
                    test_fixture="Fixture A",
                    full_test_completed=True,
                    full_test_passed=True,
                    html_content=f"<html><body><h1>Report for {board.serial_number}</h1></body></html>"
                )
                session.add(test_log)
    
    @classmethod
    def tearDownClass(cls):
        cls.db.close()
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_build_autocomplete_list(self):
        """Test building autocomplete list from database."""
        with self.db.session_scope() as session:
            autocomplete_values = set()
            
            # PIA serials
            for (serial,) in session.query(PCBABoard.serial_number).distinct():
                if serial:
                    autocomplete_values.add(serial)
            
            # PIA parts
            for (part,) in session.query(PCBABoard.part_number).distinct():
                if part:
                    autocomplete_values.add(part)
            
            # PMT serials
            for (serial,) in session.query(PMT.pmt_serial_number).distinct():
                if serial:
                    autocomplete_values.add(serial)
            
            self.assertGreater(len(autocomplete_values), 0)
            self.assertIn("ABC-001", autocomplete_values)
            self.assertIn("PART-A", autocomplete_values)
            self.assertIn("PMT-111", autocomplete_values)
        print("✓ Build autocomplete list passed")
    
    def test_02_search_by_serial(self):
        """Test searching by serial number."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            from sqlalchemy import or_
            
            term = "%ABC%"
            results = session.query(TestLog).options(
                joinedload(TestLog.pia_board),
                joinedload(TestLog.pmt_device)
            ).outerjoin(TestLog.pia_board).filter(
                PCBABoard.serial_number.ilike(term)
            ).all()
            
            self.assertEqual(len(results), 2)  # ABC-001 and ABC-002
        print("✓ Search by serial passed")
    
    def test_03_search_by_part_number(self):
        """Test searching by part number."""
        with self.db.session_scope() as session:
            from sqlalchemy.orm import joinedload
            
            term = "%PART-B%"
            results = session.query(TestLog).options(
                joinedload(TestLog.pia_board)
            ).join(TestLog.pia_board).filter(
                PCBABoard.part_number.ilike(term)
            ).all()
            
            self.assertEqual(len(results), 2)  # XYZ-001 and XYZ-002
        print("✓ Search by part number passed")
    
    def test_04_get_html_content(self):
        """Test retrieving HTML content for a test log."""
        with self.db.session_scope() as session:
            test_log = session.query(TestLog).first()
            self.assertIsNotNone(test_log)
            self.assertIsNotNone(test_log.html_content)
            self.assertIn("<html>", test_log.html_content)
        print("✓ Get HTML content passed")
    
    def test_05_compare_two_reports(self):
        """Test loading two reports for comparison."""
        with self.db.session_scope() as session:
            test_logs = session.query(TestLog).limit(2).all()
            self.assertEqual(len(test_logs), 2)
            
            # Both should have HTML content
            for tl in test_logs:
                self.assertIsNotNone(tl.html_content)
        print("✓ Compare two reports passed")


@unittest.skipUnless(OPENPYXL_AVAILABLE, "openpyxl not installed")
class TestExcelExportLogic(unittest.TestCase):
    """Test Excel export functionality."""
    
    @classmethod
    def setUpClass(cls):
        """Setup for export tests."""
        cls.temp_dir = tempfile.mkdtemp()
    
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
    
    def test_01_create_basic_excel(self):
        """Test creating a basic Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Data"
        
        # Add headers
        headers = ["Spec Name", "Measurement", "Unit", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        # Add data
        test_data = [
            ("Output Voltage 5V", 5.02, "V", "PASS"),
            ("Output Voltage 3.3V", 3.28, "V", "PASS"),
            ("Output Voltage 12V", 12.15, "V", "PASS"),
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save
        output_path = os.path.join(self.temp_dir, "test_export.xlsx")
        wb.save(output_path)
        
        self.assertTrue(os.path.exists(output_path))
        print("✓ Create basic Excel passed")
    
    def test_02_excel_with_conditional_formatting(self):
        """Test Excel with pass/fail coloring."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        
        wb = Workbook()
        ws = wb.active
        
        pass_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        fail_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        
        # Add data with conditional formatting
        data = [
            ("Spec A", 5.0, True),
            ("Spec B", 3.5, False),
            ("Spec C", 12.0, True),
        ]
        
        for row_idx, (name, value, passed) in enumerate(data, 1):
            ws.cell(row=row_idx, column=1, value=name)
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.fill = pass_fill if passed else fail_fill
        
        output_path = os.path.join(self.temp_dir, "test_conditional.xlsx")
        wb.save(output_path)
        
        self.assertTrue(os.path.exists(output_path))
        print("✓ Excel with conditional formatting passed")
    
    def test_03_excel_with_zebra_striping(self):
        """Test Excel with zebra striping."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        
        wb = Workbook()
        ws = wb.active
        
        row_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        alt_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        for row_idx in range(1, 11):
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=f"R{row_idx}C{col_idx}")
                cell.fill = alt_fill if row_idx % 2 == 0 else row_fill
        
        output_path = os.path.join(self.temp_dir, "test_zebra.xlsx")
        wb.save(output_path)
        
        self.assertTrue(os.path.exists(output_path))
        print("✓ Excel with zebra striping passed")


def run_all_tests():
    """Run all tests and print summary."""
    print("\n" + "=" * 60)
    print("PCBA Database Application - Test Suite")
    print("=" * 60 + "\n")
    
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add test classes
    suite.addTests(loader.loadTestsFromTestCase(TestDatabaseSetup))
    suite.addTests(loader.loadTestsFromTestCase(TestDatabaseQueries))
    suite.addTests(loader.loadTestsFromTestCase(TestDatabasePageLogic))
    suite.addTests(loader.loadTestsFromTestCase(TestReportsPageLogic))
    suite.addTests(loader.loadTestsFromTestCase(TestSearchPageLogic))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportLogic))
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print(f"Skipped: {len(result.skipped)}")
    
    if result.wasSuccessful():
        print("\n✅ ALL TESTS PASSED!")
    else:
        print("\n❌ SOME TESTS FAILED")
        
        if result.failures:
            print("\nFailures:")
            for test, traceback in result.failures:
                print(f"  - {test}: {traceback.split(chr(10))[0]}")
        
        if result.errors:
            print("\nErrors:")
            for test, traceback in result.errors:
                print(f"  - {test}: {traceback.split(chr(10))[0]}")
    
    print("=" * 60 + "\n")
    
    return result.wasSuccessful()


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
