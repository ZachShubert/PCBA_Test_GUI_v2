"""
Reports Page - Interactive data viewer and Excel export builder.

Features:
- View spec measurements across devices/tests for trend analysis
- Filter by device, PCBA, batch numbers
- Compare up to 10 test runs for the same device
- Customizable Excel export with:
  - Data arrangement (grouping, transposing)
  - Custom styling (colors, zebra striping, headers)
  - Pass/Fail conditional formatting
  - Embedded plot images

Author: Generated for PCBA Database Application
"""
import logging
import io
import base64
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from pathlib import Path

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QLabel, QLineEdit, QComboBox,
    QPushButton, QFrame, QScrollArea, QDateEdit, QMessageBox,
    QProgressDialog, QApplication, QFileDialog, QSplitter,
    QCheckBox, QGroupBox, QRadioButton, QButtonGroup, QSpacerItem,
    QSizePolicy, QListWidget, QListWidgetItem, QColorDialog, QSpinBox,
    QDialog, QFormLayout, QDialogButtonBox, QTabWidget, QTextEdit, QMenu
)
from PyQt6.QtCore import (
    QThread, pyqtSignal, Qt, QDate, QTimer, QSize
)
from PyQt6.QtGui import QColor, QBrush, QFont, QIcon, QPainter, QPixmap, QAction

from src.database import DatabaseManager
from src.database.database_device_tables import PCBABoard, PMT
from src.database.database_test_log_tables import TestLog, SubTest, Spec

logger = logging.getLogger(__name__)


class ExportStyle:
    """Container for export styling options."""
    
    def __init__(self):
        # Header styling
        self.header_bg_color = "#1e293b"
        self.header_text_color = "#f8fafc"
        self.header_font_bold = True
        self.header_font_size = 12
        
        # Data row styling
        self.data_bg_color = "#0f172a"
        self.data_text_color = "#f8fafc"
        self.data_font_size = 11
        
        # Zebra striping
        self.zebra_enabled = True
        self.zebra_color = "#1e293b"
        
        # Border styling
        self.border_color = "#334155"
        self.border_style = "thin"
        
        # Conditional formatting
        self.pass_color = "#22c55e"
        self.fail_color = "#ef4444"
        self.conditional_enabled = True
        
        # Layout options
        self.transpose = False
        self.include_limits = True
        self.include_units = True
        self.include_timestamps = True
        self.include_plots = True
        
        # Multiple headers
        self.group_header_enabled = True
        self.group_header_bg_color = "#334155"


class ColorButton(QPushButton):
    """A button that displays and allows selection of a color."""
    
    colorChanged = pyqtSignal(str)
    
    def __init__(self, color: str = "#ffffff", parent=None):
        super().__init__(parent)
        self._color = color
        self.setFixedSize(60, 25)
        self.clicked.connect(self._pick_color)
        self._update_style()
    
    def _update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self._color};
                border: 1px solid #334155;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                border: 2px solid #3b82f6;
            }}
        """)
    
    def _pick_color(self):
        color = QColorDialog.getColor(QColor(self._color), self)
        if color.isValid():
            self._color = color.name()
            self._update_style()
            self.colorChanged.emit(self._color)
    
    def color(self) -> str:
        return self._color
    
    def setColor(self, color: str):
        self._color = color
        self._update_style()


class ReportsPage:
    """
    Manages the reports page functionality.
    
    Provides:
    - Interactive spec data viewer with filtering
    - Trend analysis across multiple tests
    - Customizable Excel export with styling
    - Plot image embedding
    """
    
    def __init__(self, main_window, db_manager: DatabaseManager):
        """
        Initialize the reports page.
        
        Args:
            main_window: Reference to the main application window
            db_manager: Database manager instance
        """
        self.main_window = main_window
        self.db = db_manager
        
        # Current state
        self.selected_specs: List[str] = []
        self.current_data: List[Dict] = []
        self.original_data: List[Dict] = []  # Store original for reset
        self.export_style = ExportStyle()
        
        # Build the UI
        self.setup_ui()
        
        # Connect signals
        self.setup_connections()
        
        # Load initial data
        self.load_filter_options()
        
        logger.info("ReportsPage initialized")
    
    def setup_ui(self):
        """Build the reports page UI - loads from .ui file or creates programmatically."""
        from PyQt6 import uic
        from pathlib import Path
        
        mw = self.main_window
        
        # Get the reports_page widget from the UI
        if not hasattr(mw, 'reports_page'):
            logger.error("reports_page widget not found in main window")
            return
        
        reports_page = mw.reports_page
        
        # Try to load from .ui file first
        ui_file = Path(__file__).parent.parent / 'user_interfaces' / 'reports_page.ui'
        
        if ui_file.exists():
            try:
                # Clear any existing layout
                if reports_page.layout():
                    QWidget().setLayout(reports_page.layout())
                
                # Load UI
                loaded_widget = uic.loadUi(str(ui_file))
                
                # Create a layout for the reports_page and add the loaded widget
                layout = QHBoxLayout(reports_page)
                layout.setContentsMargins(0, 0, 0, 0)
                layout.setSpacing(0)
                layout.addWidget(loaded_widget)
                
                # Get references to widgets from the loaded UI
                self._get_ui_widgets(loaded_widget)
                
                # Setup color buttons (not in UI file - added dynamically)
                self._setup_color_buttons(loaded_widget)
                
                logger.info("ReportsPage UI loaded from .ui file")
                return
                
            except Exception as e:
                logger.warning(f"Failed to load reports_page.ui: {e}. Falling back to programmatic UI.")
        
        # Fallback: Create UI programmatically
        self._setup_ui_programmatic(reports_page)
    
    def _get_ui_widgets(self, ui):
        """Get widget references from loaded UI file."""
        # Sidebar controls
        self.spec_search = ui.findChild(QLineEdit, 'spec_search')
        self.spec_list = ui.findChild(QListWidget, 'spec_list')
        self.select_all_specs_btn = ui.findChild(QPushButton, 'select_all_specs_btn')
        self.clear_specs_btn = ui.findChild(QPushButton, 'clear_specs_btn')
        
        # Filters
        self.group_by_combo = ui.findChild(QComboBox, 'group_by_combo')
        self.pia_part_filter = ui.findChild(QComboBox, 'pia_part_filter')
        self.pia_serial_filter = ui.findChild(QComboBox, 'pia_serial_filter')
        self.pmt_batch_filter = ui.findChild(QComboBox, 'pmt_batch_filter')
        self.pmt_serial_filter = ui.findChild(QComboBox, 'pmt_serial_filter')
        self.fixture_filter = ui.findChild(QComboBox, 'fixture_filter')
        self.date_from = ui.findChild(QDateEdit, 'date_from')
        self.date_to = ui.findChild(QDateEdit, 'date_to')
        self.max_tests_spin = ui.findChild(QSpinBox, 'max_tests_spin')
        
        # Set default dates
        if self.date_from:
            self.date_from.setDate(QDate.currentDate().addMonths(-6))
        if self.date_to:
            self.date_to.setDate(QDate.currentDate())
        
        # Buttons
        self.generate_btn = ui.findChild(QPushButton, 'generate_btn')
        self.export_btn = ui.findChild(QPushButton, 'export_btn')
        self.export_settings_btn = ui.findChild(QPushButton, 'export_settings_btn')
        self.delete_rows_btn = ui.findChild(QPushButton, 'delete_rows_btn')
        self.delete_cols_btn = ui.findChild(QPushButton, 'delete_cols_btn')
        self.preview_style_btn = ui.findChild(QPushButton, 'preview_style_btn')
        self.reset_table_btn = ui.findChild(QPushButton, 'reset_table_btn')
        self.toggle_export_panel_btn = ui.findChild(QPushButton, 'toggle_export_panel_btn')
        
        # Labels
        self.page_title = ui.findChild(QLabel, 'page_title')
        self.page_subtitle = ui.findChild(QLabel, 'page_subtitle')
        
        # Table
        self.data_table = ui.findChild(QTableWidget, 'data_table')
        if self.data_table:
            # Enable drag-and-drop column reordering
            self.data_table.horizontalHeader().setSectionsMovable(True)
            self.data_table.horizontalHeader().setDragEnabled(True)
            self.data_table.horizontalHeader().setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
            self.data_table.verticalHeader().setVisible(True)
        
        # Export panel
        self.export_panel = ui.findChild(QFrame, 'export_panel')
        self.export_tabs = ui.findChild(QTabWidget, 'export_tabs')
        
        # Layout tab checkboxes
        self.transpose_checkbox = ui.findChild(QCheckBox, 'transpose_checkbox')
        self.include_limits_checkbox = ui.findChild(QCheckBox, 'include_limits_checkbox')
        self.include_units_checkbox = ui.findChild(QCheckBox, 'include_units_checkbox')
        self.include_timestamps_checkbox = ui.findChild(QCheckBox, 'include_timestamps_checkbox')
        self.include_plots_checkbox = ui.findChild(QCheckBox, 'include_plots_checkbox')
        self.group_headers_checkbox = ui.findChild(QCheckBox, 'group_headers_checkbox')
        
        # Colors tab
        self.zebra_checkbox = ui.findChild(QCheckBox, 'zebra_checkbox')
        
        # Conditional tab
        self.cond_format_checkbox = ui.findChild(QCheckBox, 'cond_format_checkbox')
    
    def _setup_color_buttons(self, ui):
        """Setup color picker buttons (added dynamically since custom widgets aren't in .ui)."""
        # Find the colors tab
        colors_tab = ui.findChild(QWidget, 'colors_tab')
        if not colors_tab:
            return
        
        # Get or create layout
        colors_layout = colors_tab.layout()
        if not colors_layout:
            return
        
        # Remove placeholder label if it exists
        info_label = ui.findChild(QLabel, 'colors_info_label')
        if info_label:
            info_label.setVisible(False)
        
        # Create color button grid
        color_grid = QHBoxLayout()
        
        # Header colors
        header_group = QVBoxLayout()
        header_group.addWidget(QLabel("Header"))
        self.header_bg_btn = ColorButton("#1e293b", "Background")
        self.header_text_btn = ColorButton("#f8fafc", "Text")
        header_group.addWidget(self.header_bg_btn)
        header_group.addWidget(self.header_text_btn)
        color_grid.addLayout(header_group)
        
        # Data colors
        data_group = QVBoxLayout()
        data_group.addWidget(QLabel("Data"))
        self.data_bg_btn = ColorButton("#0f172a", "Background")
        self.data_text_btn = ColorButton("#f8fafc", "Text")
        data_group.addWidget(self.data_bg_btn)
        data_group.addWidget(self.data_text_btn)
        color_grid.addLayout(data_group)
        
        # Other colors
        other_group = QVBoxLayout()
        other_group.addWidget(QLabel("Other"))
        self.zebra_color_btn = ColorButton("#1e293b", "Zebra")
        self.border_color_btn = ColorButton("#334155", "Border")
        other_group.addWidget(self.zebra_color_btn)
        other_group.addWidget(self.border_color_btn)
        color_grid.addLayout(other_group)
        
        # Pass/Fail colors
        pf_group = QVBoxLayout()
        pf_group.addWidget(QLabel("Pass/Fail"))
        self.pass_color_btn = ColorButton("#22c55e", "Pass")
        self.fail_color_btn = ColorButton("#ef4444", "Fail")
        pf_group.addWidget(self.pass_color_btn)
        pf_group.addWidget(self.fail_color_btn)
        color_grid.addLayout(pf_group)
        
        # Insert at beginning of layout (after removing placeholder)
        colors_layout.insertLayout(0, color_grid)
    
    def _setup_ui_programmatic(self, reports_page):
        """Fallback: Create UI programmatically if .ui file not available."""
        # Clear any existing layout
        if reports_page.layout():
            QWidget().setLayout(reports_page.layout())
        
        # Main horizontal layout
        main_layout = QHBoxLayout(reports_page)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ==================== SIDEBAR ====================
        sidebar_frame = QFrame()
        sidebar_frame.setProperty('class', 'side-bar-frame')
        sidebar_frame.setMinimumWidth(300)
        sidebar_frame.setMaximumWidth(300)
        
        sidebar_layout = QVBoxLayout(sidebar_frame)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)
        
        # Sidebar scroll area
        sidebar_scroll = QScrollArea()
        sidebar_scroll.setWidgetResizable(True)
        sidebar_scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        sidebar_content = QWidget()
        sidebar_content_layout = QVBoxLayout(sidebar_content)
        sidebar_content_layout.setContentsMargins(15, 15, 15, 15)
        sidebar_content_layout.setSpacing(12)
        
        # --- Spec Selection Section ---
        spec_label = QLabel("Spec Selection")
        spec_label.setProperty('class', 'heading-3')
        sidebar_content_layout.addWidget(spec_label)
        
        # Spec search
        self.spec_search = QLineEdit()
        self.spec_search.setPlaceholderText("Search specs...")
        self.spec_search.setClearButtonEnabled(True)
        sidebar_content_layout.addWidget(self.spec_search)
        
        # Spec list (multi-select)
        self.spec_list = QListWidget()
        self.spec_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.spec_list.setMaximumHeight(200)
        sidebar_content_layout.addWidget(self.spec_list)
        
        # Select all / Clear buttons
        spec_btn_layout = QHBoxLayout()
        self.select_all_specs_btn = QPushButton("Select All")
        self.select_all_specs_btn.setProperty('class', 'btn-ghost btn-sm')
        spec_btn_layout.addWidget(self.select_all_specs_btn)
        
        self.clear_specs_btn = QPushButton("Clear")
        self.clear_specs_btn.setProperty('class', 'btn-ghost btn-sm')
        spec_btn_layout.addWidget(self.clear_specs_btn)
        sidebar_content_layout.addLayout(spec_btn_layout)
        
        # Divider
        self._add_divider(sidebar_content_layout)
        
        # --- Filter Section ---
        filter_label = QLabel("Filters")
        filter_label.setProperty('class', 'heading-3')
        sidebar_content_layout.addWidget(filter_label)
        
        # Group by selector
        sidebar_content_layout.addWidget(QLabel("Group By"))
        self.group_by_combo = QComboBox()
        self.group_by_combo.addItems([
            "None",
            "PIA Part Number",
            "PIA Serial Number",
            "PMT Batch Number",
            "PMT Serial Number",
            "Test Name",
            "Test Fixture"
        ])
        sidebar_content_layout.addWidget(self.group_by_combo)
        
        # PIA Part Number filter
        sidebar_content_layout.addWidget(QLabel("PIA Part Number"))
        self.pia_part_filter = QComboBox()
        self.pia_part_filter.addItem("All", None)
        sidebar_content_layout.addWidget(self.pia_part_filter)
        
        # PIA Serial Number filter
        sidebar_content_layout.addWidget(QLabel("PIA Serial Number"))
        self.pia_serial_filter = QComboBox()
        self.pia_serial_filter.addItem("All", None)
        self.pia_serial_filter.setEditable(True)
        self.pia_serial_filter.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        sidebar_content_layout.addWidget(self.pia_serial_filter)
        
        # PMT Batch filter
        sidebar_content_layout.addWidget(QLabel("PMT Batch Number"))
        self.pmt_batch_filter = QComboBox()
        self.pmt_batch_filter.addItem("All", None)
        sidebar_content_layout.addWidget(self.pmt_batch_filter)
        
        # PMT Serial filter
        sidebar_content_layout.addWidget(QLabel("PMT Serial Number"))
        self.pmt_serial_filter = QComboBox()
        self.pmt_serial_filter.addItem("All", None)
        self.pmt_serial_filter.setEditable(True)
        self.pmt_serial_filter.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        sidebar_content_layout.addWidget(self.pmt_serial_filter)
        
        # Test Fixture filter
        sidebar_content_layout.addWidget(QLabel("Test Fixture"))
        self.fixture_filter = QComboBox()
        self.fixture_filter.addItem("All", None)
        sidebar_content_layout.addWidget(self.fixture_filter)
        
        # Date range
        sidebar_content_layout.addWidget(QLabel("Date Range"))
        date_layout = QHBoxLayout()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addMonths(-6))
        self.date_from.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.date_from)
        
        date_layout.addWidget(QLabel("to"))
        
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setDisplayFormat("yyyy-MM-dd")
        date_layout.addWidget(self.date_to)
        sidebar_content_layout.addLayout(date_layout)
        
        # Max tests per device
        max_tests_layout = QHBoxLayout()
        max_tests_layout.addWidget(QLabel("Max Tests/Device"))
        self.max_tests_spin = QSpinBox()
        self.max_tests_spin.setRange(1, 10)
        self.max_tests_spin.setValue(10)
        max_tests_layout.addWidget(self.max_tests_spin)
        sidebar_content_layout.addLayout(max_tests_layout)
        
        # Divider
        self._add_divider(sidebar_content_layout)
        
        # Generate Report button
        self.generate_btn = QPushButton("📊 Generate Report")
        self.generate_btn.setProperty('class', 'btn-primary')
        sidebar_content_layout.addWidget(self.generate_btn)
        
        # Spacer
        sidebar_content_layout.addStretch()
        
        sidebar_scroll.setWidget(sidebar_content)
        sidebar_layout.addWidget(sidebar_scroll)
        
        main_layout.addWidget(sidebar_frame)
        
        # ==================== MAIN CONTENT AREA ====================
        content_frame = QFrame()
        content_layout = QVBoxLayout(content_frame)
        content_layout.setContentsMargins(20, 15, 20, 15)
        content_layout.setSpacing(15)
        
        # --- Header Section ---
        header_layout = QHBoxLayout()
        
        # Title and subtitle
        title_layout = QVBoxLayout()
        title_layout.setSpacing(2)
        
        self.page_title = QLabel("Reports & Export")
        self.page_title.setProperty('class', 'heading-1')
        title_layout.addWidget(self.page_title)
        
        self.page_subtitle = QLabel("Select specs and filters to generate a report")
        self.page_subtitle.setStyleSheet("color: #94a3b8;")
        title_layout.addWidget(self.page_subtitle)
        
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        
        # Export button
        self.export_btn = QPushButton("📥 Export to Excel")
        self.export_btn.setProperty('class', 'btn-primary')
        self.export_btn.setEnabled(False)
        header_layout.addWidget(self.export_btn)
        
        # Export settings button
        self.export_settings_btn = QPushButton("⚙️")
        self.export_settings_btn.setProperty('class', 'btn-icon')
        self.export_settings_btn.setToolTip("Export Settings")
        header_layout.addWidget(self.export_settings_btn)
        
        content_layout.addLayout(header_layout)
        
        # --- Table Controls ---
        table_controls = QHBoxLayout()
        
        self.delete_rows_btn = QPushButton("🗑️ Delete Selected Rows")
        self.delete_rows_btn.setProperty('class', 'btn-danger btn-sm')
        self.delete_rows_btn.setEnabled(False)
        self.delete_rows_btn.setToolTip("Delete selected rows from the report")
        table_controls.addWidget(self.delete_rows_btn)
        
        self.delete_cols_btn = QPushButton("🗑️ Delete Selected Columns")
        self.delete_cols_btn.setProperty('class', 'btn-danger btn-sm')
        self.delete_cols_btn.setEnabled(False)
        self.delete_cols_btn.setToolTip("Delete selected columns from the report")
        table_controls.addWidget(self.delete_cols_btn)
        
        table_controls.addStretch()
        
        self.preview_style_btn = QPushButton("👁️ Preview Style")
        self.preview_style_btn.setProperty('class', 'btn-info btn-sm')
        self.preview_style_btn.setToolTip("Preview export styling in the table")
        self.preview_style_btn.setCheckable(True)
        table_controls.addWidget(self.preview_style_btn)
        
        self.reset_table_btn = QPushButton("↻ Reset Table")
        self.reset_table_btn.setProperty('class', 'btn-ghost btn-sm')
        self.reset_table_btn.setToolTip("Reset table to original generated data")
        table_controls.addWidget(self.reset_table_btn)
        
        content_layout.addLayout(table_controls)
        
        # --- Main Table ---
        self.data_table = QTableWidget()
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.data_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setSortingEnabled(True)
        self.data_table.setShowGrid(True)
        self.data_table.verticalHeader().setVisible(True)
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Enable drag-and-drop column reordering
        self.data_table.horizontalHeader().setSectionsMovable(True)
        self.data_table.horizontalHeader().setDragEnabled(True)
        self.data_table.horizontalHeader().setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        
        # Context menu for table
        self.data_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        
        content_layout.addWidget(self.data_table, stretch=1)
        
        # --- Export Configuration Panel ---
        self.export_panel = QFrame()
        self.export_panel.setProperty('class', 'container-secondary')
        self.export_panel.setMinimumHeight(220)
        self.export_panel.setMaximumHeight(280)
        
        export_layout = QVBoxLayout(self.export_panel)
        export_layout.setContentsMargins(15, 15, 15, 15)
        export_layout.setSpacing(10)
        
        # Export panel header
        export_header = QHBoxLayout()
        export_title = QLabel("📋 Export Configuration")
        export_title.setProperty('class', 'heading-3')
        export_header.addWidget(export_title)
        export_header.addStretch()
        
        # Toggle panel visibility
        self.toggle_export_panel_btn = QPushButton("▼")
        self.toggle_export_panel_btn.setProperty('class', 'btn-ghost btn-sm')
        self.toggle_export_panel_btn.setFixedWidth(30)
        export_header.addWidget(self.toggle_export_panel_btn)
        
        export_layout.addLayout(export_header)
        
        # Export options tabs
        self.export_tabs = QTabWidget()
        self.export_tabs.setMaximumHeight(180)
        
        # --- Layout Tab ---
        layout_tab = QWidget()
        layout_tab_layout = QVBoxLayout(layout_tab)
        layout_tab_layout.setContentsMargins(10, 10, 10, 10)
        layout_tab_layout.setSpacing(8)
        
        # Transpose option
        self.transpose_checkbox = QCheckBox("Transpose Data (Rows ↔ Columns)")
        layout_tab_layout.addWidget(self.transpose_checkbox)
        
        # Include options
        include_layout = QHBoxLayout()
        self.include_limits_checkbox = QCheckBox("Include Limits")
        self.include_limits_checkbox.setChecked(True)
        include_layout.addWidget(self.include_limits_checkbox)
        
        self.include_units_checkbox = QCheckBox("Include Units")
        self.include_units_checkbox.setChecked(True)
        include_layout.addWidget(self.include_units_checkbox)
        
        self.include_timestamps_checkbox = QCheckBox("Include Timestamps")
        self.include_timestamps_checkbox.setChecked(True)
        include_layout.addWidget(self.include_timestamps_checkbox)
        layout_tab_layout.addLayout(include_layout)
        
        # Plot options
        plot_layout = QHBoxLayout()
        self.include_plots_checkbox = QCheckBox("Include Plot Images")
        self.include_plots_checkbox.setChecked(True)
        plot_layout.addWidget(self.include_plots_checkbox)
        
        self.group_headers_checkbox = QCheckBox("Group Headers")
        self.group_headers_checkbox.setChecked(True)
        plot_layout.addWidget(self.group_headers_checkbox)
        layout_tab_layout.addLayout(plot_layout)
        
        layout_tab_layout.addStretch()
        self.export_tabs.addTab(layout_tab, "Layout")
        
        # --- Colors Tab ---
        colors_tab = QWidget()
        colors_tab_layout = QVBoxLayout(colors_tab)
        colors_tab_layout.setContentsMargins(10, 10, 10, 10)
        colors_tab_layout.setSpacing(8)
        
        # Header colors
        header_colors = QHBoxLayout()
        header_colors.addWidget(QLabel("Header Background:"))
        self.header_bg_btn = ColorButton(self.export_style.header_bg_color)
        header_colors.addWidget(self.header_bg_btn)
        
        header_colors.addWidget(QLabel("Text:"))
        self.header_text_btn = ColorButton(self.export_style.header_text_color)
        header_colors.addWidget(self.header_text_btn)
        header_colors.addStretch()
        colors_tab_layout.addLayout(header_colors)
        
        # Data colors
        data_colors = QHBoxLayout()
        data_colors.addWidget(QLabel("Data Background:"))
        self.data_bg_btn = ColorButton(self.export_style.data_bg_color)
        data_colors.addWidget(self.data_bg_btn)
        
        data_colors.addWidget(QLabel("Text:"))
        self.data_text_btn = ColorButton(self.export_style.data_text_color)
        data_colors.addWidget(self.data_text_btn)
        data_colors.addStretch()
        colors_tab_layout.addLayout(data_colors)
        
        # Zebra striping
        zebra_layout = QHBoxLayout()
        self.zebra_checkbox = QCheckBox("Zebra Striping")
        self.zebra_checkbox.setChecked(True)
        zebra_layout.addWidget(self.zebra_checkbox)
        
        zebra_layout.addWidget(QLabel("Alt Row:"))
        self.zebra_color_btn = ColorButton(self.export_style.zebra_color)
        zebra_layout.addWidget(self.zebra_color_btn)
        zebra_layout.addStretch()
        colors_tab_layout.addLayout(zebra_layout)
        
        # Border color
        border_layout = QHBoxLayout()
        border_layout.addWidget(QLabel("Border Color:"))
        self.border_color_btn = ColorButton(self.export_style.border_color)
        border_layout.addWidget(self.border_color_btn)
        border_layout.addStretch()
        colors_tab_layout.addLayout(border_layout)
        
        colors_tab_layout.addStretch()
        self.export_tabs.addTab(colors_tab, "Colors")
        
        # --- Conditional Formatting Tab ---
        cond_tab = QWidget()
        cond_tab_layout = QVBoxLayout(cond_tab)
        cond_tab_layout.setContentsMargins(10, 10, 10, 10)
        cond_tab_layout.setSpacing(8)
        
        # Enable conditional formatting
        self.cond_format_checkbox = QCheckBox("Enable Pass/Fail Conditional Formatting")
        self.cond_format_checkbox.setChecked(True)
        cond_tab_layout.addWidget(self.cond_format_checkbox)
        
        # Pass/Fail colors
        pf_colors = QHBoxLayout()
        pf_colors.addWidget(QLabel("Pass Color:"))
        self.pass_color_btn = ColorButton(self.export_style.pass_color)
        pf_colors.addWidget(self.pass_color_btn)
        
        pf_colors.addWidget(QLabel("Fail Color:"))
        self.fail_color_btn = ColorButton(self.export_style.fail_color)
        pf_colors.addWidget(self.fail_color_btn)
        pf_colors.addStretch()
        cond_tab_layout.addLayout(pf_colors)
        
        # Description
        cond_desc = QLabel("Values within limits will be highlighted green,\nvalues outside limits will be highlighted red.")
        cond_desc.setStyleSheet("color: #64748b; font-size: 11px;")
        cond_tab_layout.addWidget(cond_desc)
        
        cond_tab_layout.addStretch()
        self.export_tabs.addTab(cond_tab, "Conditional")
        
        export_layout.addWidget(self.export_tabs)
        
        content_layout.addWidget(self.export_panel)
        
        main_layout.addWidget(content_frame, stretch=1)
        
        # Store widget references
        mw.reports_data_table = self.data_table
        mw.reports_export_panel = self.export_panel
    
    def _add_divider(self, layout: QVBoxLayout):
        """Add a horizontal divider line to a layout."""
        divider = QFrame()
        divider.setMinimumHeight(2)
        divider.setMaximumHeight(2)
        divider.setStyleSheet("background-color: #334155;")
        layout.addWidget(divider)
    
    def setup_connections(self):
        """Connect UI signals to handlers."""
        # Spec selection
        self.spec_search.textChanged.connect(self.filter_spec_list)
        self.select_all_specs_btn.clicked.connect(self.select_all_specs)
        self.clear_specs_btn.clicked.connect(self.clear_spec_selection)
        
        # Generate report
        self.generate_btn.clicked.connect(self.generate_report)
        
        # Export
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_settings_btn.clicked.connect(self.show_export_settings)
        
        # Toggle export panel
        self.toggle_export_panel_btn.clicked.connect(self.toggle_export_panel)
        
        # Color button changes
        self.header_bg_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'header_bg_color', c))
        self.header_text_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'header_text_color', c))
        self.data_bg_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'data_bg_color', c))
        self.data_text_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'data_text_color', c))
        self.zebra_color_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'zebra_color', c))
        self.border_color_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'border_color', c))
        self.pass_color_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'pass_color', c))
        self.fail_color_btn.colorChanged.connect(lambda c: setattr(self.export_style, 'fail_color', c))
        
        # Table control buttons
        self.delete_rows_btn.clicked.connect(self.delete_selected_rows)
        self.delete_cols_btn.clicked.connect(self.delete_selected_columns)
        self.preview_style_btn.toggled.connect(self.toggle_style_preview)
        self.reset_table_btn.clicked.connect(self.reset_table_data)
        
        # Table selection changed
        self.data_table.itemSelectionChanged.connect(self.on_table_selection_changed)
        
        # Table context menu
        self.data_table.customContextMenuRequested.connect(self.show_table_context_menu)
        
        # Auto-update preview when style options change
        self.zebra_checkbox.toggled.connect(self._on_style_option_changed)
        self.cond_format_checkbox.toggled.connect(self._on_style_option_changed)
        
        logger.info("ReportsPage connections established")
    
    def load_filter_options(self):
        """Load filter options from database."""
        try:
            with self.db.session_scope() as session:
                # Load spec names
                specs = session.query(Spec.name).distinct().order_by(Spec.name).all()
                self.spec_list.clear()
                for (spec_name,) in specs:
                    if spec_name:
                        item = QListWidgetItem(spec_name)
                        self.spec_list.addItem(item)
                
                # Load PIA part numbers
                parts = session.query(PCBABoard.part_number).distinct().order_by(PCBABoard.part_number).all()
                self.pia_part_filter.clear()
                self.pia_part_filter.addItem("All", None)
                for (part,) in parts:
                    if part:
                        self.pia_part_filter.addItem(part, part)
                
                # Load PIA serial numbers
                serials = session.query(PCBABoard.serial_number).distinct().order_by(PCBABoard.serial_number).all()
                self.pia_serial_filter.clear()
                self.pia_serial_filter.addItem("All", None)
                for (serial,) in serials:
                    if serial:
                        self.pia_serial_filter.addItem(serial, serial)
                
                # Load PMT batch numbers
                batches = session.query(PMT.batch_number).distinct().order_by(PMT.batch_number).all()
                self.pmt_batch_filter.clear()
                self.pmt_batch_filter.addItem("All", None)
                for (batch,) in batches:
                    if batch:
                        self.pmt_batch_filter.addItem(batch, batch)
                
                # Load PMT serial numbers
                pmt_serials = session.query(PMT.pmt_serial_number).distinct().order_by(PMT.pmt_serial_number).all()
                self.pmt_serial_filter.clear()
                self.pmt_serial_filter.addItem("All", None)
                for (serial,) in pmt_serials:
                    if serial:
                        self.pmt_serial_filter.addItem(serial, serial)
                
                # Load test fixtures
                fixtures = session.query(TestLog.test_fixture).distinct().order_by(TestLog.test_fixture).all()
                self.fixture_filter.clear()
                self.fixture_filter.addItem("All", None)
                for (fixture,) in fixtures:
                    if fixture:
                        self.fixture_filter.addItem(fixture, fixture)
                
                logger.info(f"Loaded {self.spec_list.count()} specs into filter")
                
        except Exception as e:
            logger.exception("Error loading filter options")
    
    def filter_spec_list(self, text: str):
        """Filter the spec list based on search text."""
        search_text = text.lower()
        for i in range(self.spec_list.count()):
            item = self.spec_list.item(i)
            item.setHidden(search_text not in item.text().lower())
    
    def select_all_specs(self):
        """Select all visible specs in the list."""
        for i in range(self.spec_list.count()):
            item = self.spec_list.item(i)
            if not item.isHidden():
                item.setSelected(True)
    
    def clear_spec_selection(self):
        """Clear all spec selections."""
        self.spec_list.clearSelection()
    
    def toggle_export_panel(self):
        """Toggle the export configuration panel visibility."""
        if self.export_panel.maximumHeight() > 50:
            self.export_panel.setMaximumHeight(50)
            self.toggle_export_panel_btn.setText("▶")
        else:
            self.export_panel.setMaximumHeight(280)
            self.toggle_export_panel_btn.setText("▼")
    
    def on_table_selection_changed(self):
        """Handle table selection changes to enable/disable delete buttons."""
        selected_items = self.data_table.selectedItems()
        has_selection = len(selected_items) > 0
        
        # Enable delete rows if any items are selected
        self.delete_rows_btn.setEnabled(has_selection)
        
        # Enable delete columns if any items are selected
        self.delete_cols_btn.setEnabled(has_selection)
    
    def delete_selected_rows(self):
        """Delete selected rows from the table."""
        selected_rows = set()
        for item in self.data_table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            return
        
        # Confirm deletion
        reply = QMessageBox.question(
            self.main_window,
            "Delete Rows",
            f"Delete {len(selected_rows)} selected row(s)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Delete rows from bottom to top to preserve indices
        for row in sorted(selected_rows, reverse=True):
            self.data_table.removeRow(row)
            # Also remove from current_data
            if row < len(self.current_data):
                del self.current_data[row]
        
        self.page_subtitle.setText(f"{self.data_table.rowCount()} measurements")
    
    def delete_selected_columns(self):
        """Delete selected columns from the table."""
        selected_cols = set()
        for item in self.data_table.selectedItems():
            selected_cols.add(item.column())
        
        if not selected_cols:
            return
        
        # Get column names for confirmation
        col_names = []
        for col in selected_cols:
            header = self.data_table.horizontalHeaderItem(col)
            if header:
                col_names.append(header.text())
        
        reply = QMessageBox.question(
            self.main_window,
            "Delete Columns",
            f"Delete {len(selected_cols)} column(s)?\n\n{', '.join(col_names)}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Delete columns from right to left to preserve indices
        for col in sorted(selected_cols, reverse=True):
            self.data_table.removeColumn(col)
    
    def reset_table_data(self):
        """Reset table to original generated data."""
        if not self.original_data:
            QMessageBox.information(
                self.main_window,
                "No Data",
                "No original data to reset to. Please generate a report first."
            )
            return
        
        reply = QMessageBox.question(
            self.main_window,
            "Reset Table",
            "Reset table to original generated data?\n\nThis will undo any row/column deletions.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            import copy
            self.current_data = copy.deepcopy(self.original_data)
            self._populate_data_table()
            self.page_subtitle.setText(f"{len(self.current_data)} measurements")
    
    def toggle_style_preview(self, enabled: bool):
        """Toggle style preview in the table."""
        if enabled:
            self._apply_style_preview()
        else:
            # Reset to default styling
            self._populate_data_table()
    
    def _on_style_option_changed(self):
        """Handle style option changes - update preview if active."""
        if self.preview_style_btn.isChecked():
            self._apply_style_preview()
    
    def _apply_style_preview(self):
        """Apply export styling to the table as a preview."""
        self._update_export_style()
        style = self.export_style
        
        # Convert hex to QColor
        def hex_to_qcolor(hex_color: str) -> QColor:
            return QColor(hex_color)
        
        header_bg = hex_to_qcolor(style.header_bg_color)
        header_text = hex_to_qcolor(style.header_text_color)
        data_bg = hex_to_qcolor(style.data_bg_color)
        data_text = hex_to_qcolor(style.data_text_color)
        zebra_bg = hex_to_qcolor(style.zebra_color)
        pass_bg = hex_to_qcolor(style.pass_color)
        fail_bg = hex_to_qcolor(style.fail_color)
        
        # Apply header styling
        header = self.data_table.horizontalHeader()
        header.setStyleSheet(f"""
            QHeaderView::section {{
                background-color: {style.header_bg_color};
                color: {style.header_text_color};
                font-weight: bold;
                padding: 8px;
                border: 1px solid {style.border_color};
            }}
        """)
        
        # Apply cell styling
        for row in range(self.data_table.rowCount()):
            # Determine row background
            if style.zebra_enabled and row % 2 == 1:
                row_bg = zebra_bg
            else:
                row_bg = data_bg
            
            for col in range(self.data_table.columnCount()):
                item = self.data_table.item(row, col)
                if not item:
                    continue
                
                # Apply background
                cell_bg = row_bg
                
                # Apply conditional formatting for Status and Measurement columns
                if style.conditional_enabled:
                    header_item = self.data_table.horizontalHeaderItem(col)
                    if header_item:
                        header_text_val = header_item.text()
                        if header_text_val in ("Status", "Measurement"):
                            # Check if this row has pass/fail data
                            if row < len(self.current_data):
                                passed = self.current_data[row].get('passed')
                                if passed is True:
                                    cell_bg = pass_bg
                                elif passed is False:
                                    cell_bg = fail_bg
                
                item.setBackground(QBrush(cell_bg))
                item.setForeground(QBrush(data_text))
    
    def show_table_context_menu(self, position):
        """Show context menu for table operations."""
        menu = QMenu()
        
        delete_rows_action = QAction("Delete Selected Rows", self.main_window)
        delete_rows_action.triggered.connect(self.delete_selected_rows)
        menu.addAction(delete_rows_action)
        
        delete_cols_action = QAction("Delete Selected Columns", self.main_window)
        delete_cols_action.triggered.connect(self.delete_selected_columns)
        menu.addAction(delete_cols_action)
        
        menu.addSeparator()
        
        reset_action = QAction("Reset Table", self.main_window)
        reset_action.triggered.connect(self.reset_table_data)
        menu.addAction(reset_action)
        
        menu.exec(self.data_table.viewport().mapToGlobal(position))
    
    def get_selected_specs(self) -> List[str]:
        """Get list of selected spec names."""
        return [item.text() for item in self.spec_list.selectedItems()]
    
    def generate_report(self):
        """Generate report based on selected specs and filters."""
        selected_specs = self.get_selected_specs()
        
        if not selected_specs:
            QMessageBox.warning(
                self.main_window,
                "No Specs Selected",
                "Please select at least one spec to generate a report."
            )
            return
        
        try:
            # Show progress
            progress = QProgressDialog(
                "Generating report...", "Cancel", 0, 100, self.main_window
            )
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.show()
            QApplication.processEvents()
            
            # Build query with filters
            with self.db.session_scope() as session:
                from sqlalchemy.orm import joinedload
                from sqlalchemy import desc, and_, or_
                
                progress.setValue(10)
                QApplication.processEvents()
                
                # Base query
                query = session.query(Spec).options(
                    joinedload(Spec.sub_test).joinedload(SubTest.test_log).joinedload(TestLog.pia_board),
                    joinedload(Spec.sub_test).joinedload(SubTest.test_log).joinedload(TestLog.pmt_device)
                ).join(Spec.sub_test).join(SubTest.test_log)
                
                # Filter by selected specs
                query = query.filter(Spec.name.in_(selected_specs))
                
                # Apply filters
                pia_part = self.pia_part_filter.currentData()
                if pia_part:
                    query = query.join(TestLog.pia_board).filter(PCBABoard.part_number == pia_part)
                
                pia_serial = self.pia_serial_filter.currentData()
                if pia_serial:
                    query = query.join(TestLog.pia_board, isouter=True).filter(PCBABoard.serial_number == pia_serial)
                
                pmt_batch = self.pmt_batch_filter.currentData()
                if pmt_batch:
                    query = query.join(TestLog.pmt_device, isouter=True).filter(PMT.batch_number == pmt_batch)
                
                pmt_serial = self.pmt_serial_filter.currentData()
                if pmt_serial:
                    query = query.join(TestLog.pmt_device, isouter=True).filter(PMT.pmt_serial_number == pmt_serial)
                
                fixture = self.fixture_filter.currentData()
                if fixture:
                    query = query.filter(TestLog.test_fixture == fixture)
                
                # Date range
                from_date = self.date_from.date().toPyDate()
                to_date = self.date_to.date().toPyDate()
                query = query.filter(
                    TestLog.created_at >= datetime.combine(from_date, datetime.min.time()),
                    TestLog.created_at <= datetime.combine(to_date, datetime.max.time())
                )
                
                # Order by date
                query = query.order_by(desc(TestLog.created_at))
                
                progress.setValue(30)
                QApplication.processEvents()
                
                # Execute query
                specs = query.all()
                
                progress.setValue(50)
                QApplication.processEvents()
                
                # Process results into display format
                self.current_data = self._process_spec_data(specs, session)
                
                # Store original data for reset functionality
                import copy
                self.original_data = copy.deepcopy(self.current_data)
                
                progress.setValue(80)
                QApplication.processEvents()
                
                # Populate table
                self._populate_data_table()
                
                progress.setValue(100)
                
                # Update UI
                self.export_btn.setEnabled(len(self.current_data) > 0)
                self.page_subtitle.setText(f"{len(self.current_data)} measurements from {len(selected_specs)} specs")
                
                logger.info(f"Generated report with {len(self.current_data)} rows")
                
        except Exception as e:
            logger.exception("Error generating report")
            QMessageBox.critical(
                self.main_window,
                "Error",
                f"Failed to generate report: {str(e)}"
            )
        finally:
            progress.close()
    
    def _process_spec_data(self, specs: List[Spec], session) -> List[Dict]:
        """Process spec data into a format suitable for display and export."""
        data = []
        max_tests = self.max_tests_spin.value()
        
        # Group specs by device + spec name
        device_spec_groups: Dict[Tuple, List[Spec]] = {}
        
        for spec in specs:
            if not spec.sub_test or not spec.sub_test.test_log:
                continue
            
            test_log = spec.sub_test.test_log
            
            # Create device key
            pia_serial = test_log.pia_board.serial_number if test_log.pia_board else "N/A"
            pmt_serial = test_log.pmt_device.pmt_serial_number if test_log.pmt_device else "N/A"
            
            key = (pia_serial, pmt_serial, spec.name)
            
            if key not in device_spec_groups:
                device_spec_groups[key] = []
            
            # Only keep up to max_tests per device
            if len(device_spec_groups[key]) < max_tests:
                device_spec_groups[key].append(spec)
        
        # Convert to flat data structure
        for (pia_serial, pmt_serial, spec_name), spec_list in device_spec_groups.items():
            for spec in spec_list:
                test_log = spec.sub_test.test_log
                
                # Determine pass/fail
                passed = None
                if spec.measurement is not None:
                    if spec.lower_limit is not None and spec.upper_limit is not None:
                        passed = spec.lower_limit <= spec.measurement <= spec.upper_limit
                    elif spec.lower_limit is not None:
                        passed = spec.measurement >= spec.lower_limit
                    elif spec.upper_limit is not None:
                        passed = spec.measurement <= spec.upper_limit
                
                row = {
                    'spec_name': spec.name,
                    'measurement': spec.measurement,
                    'unit': spec.unit or '',
                    'lower_limit': spec.lower_limit,
                    'upper_limit': spec.upper_limit,
                    'nominal': spec.nominal,
                    'result': spec.result,
                    'passed': passed,
                    'has_plot': spec.has_plot,
                    'plot_data': spec.plot_data,
                    'plot_image': spec.plot_image,
                    'pia_serial': pia_serial,
                    'pia_part': test_log.pia_board.part_number if test_log.pia_board else 'N/A',
                    'pmt_serial': pmt_serial,
                    'pmt_batch': test_log.pmt_device.batch_number if test_log.pmt_device else 'N/A',
                    'test_name': test_log.name or 'N/A',
                    'test_fixture': test_log.test_fixture or 'N/A',
                    'test_date': test_log.created_at,
                    'test_log_id': test_log.id,
                    '_spec': spec,
                }
                data.append(row)
        
        return data
    
    def _populate_data_table(self):
        """Populate the data table with current data."""
        columns = [
            "Spec Name", "Measurement", "Unit", "Lower Limit", "Upper Limit",
            "Status", "PIA Serial", "PIA Part", "PMT Serial", "PMT Batch",
            "Test Name", "Test Fixture", "Test Date"
        ]
        
        self.data_table.clear()
        self.data_table.setRowCount(len(self.current_data))
        self.data_table.setColumnCount(len(columns))
        self.data_table.setHorizontalHeaderLabels(columns)
        
        for row_idx, row_data in enumerate(self.current_data):
            # Spec Name
            item = QTableWidgetItem(row_data['spec_name'])
            item.setData(Qt.ItemDataRole.UserRole, row_data)
            self.data_table.setItem(row_idx, 0, item)
            
            # Measurement
            meas = row_data['measurement']
            meas_str = f"{meas:.6g}" if meas is not None else "N/A"
            meas_item = QTableWidgetItem(meas_str)
            
            # Apply pass/fail coloring to measurement
            if row_data['passed'] is True:
                meas_item.setForeground(QBrush(QColor('#22c55e')))
            elif row_data['passed'] is False:
                meas_item.setForeground(QBrush(QColor('#ef4444')))
            
            self.data_table.setItem(row_idx, 1, meas_item)
            
            # Unit
            self.data_table.setItem(row_idx, 2, QTableWidgetItem(row_data['unit']))
            
            # Lower Limit
            lower = row_data['lower_limit']
            self.data_table.setItem(row_idx, 3, QTableWidgetItem(f"{lower:.6g}" if lower is not None else "N/A"))
            
            # Upper Limit
            upper = row_data['upper_limit']
            self.data_table.setItem(row_idx, 4, QTableWidgetItem(f"{upper:.6g}" if upper is not None else "N/A"))
            
            # Status
            if row_data['passed'] is True:
                status_item = QTableWidgetItem("✓ PASS")
                status_item.setForeground(QBrush(QColor('#22c55e')))
            elif row_data['passed'] is False:
                status_item = QTableWidgetItem("✗ FAIL")
                status_item.setForeground(QBrush(QColor('#ef4444')))
            else:
                status_item = QTableWidgetItem("N/A")
            status_item.setFont(QFont('', -1, QFont.Weight.Bold))
            self.data_table.setItem(row_idx, 5, status_item)
            
            # Device info
            self.data_table.setItem(row_idx, 6, QTableWidgetItem(row_data['pia_serial']))
            self.data_table.setItem(row_idx, 7, QTableWidgetItem(row_data['pia_part']))
            self.data_table.setItem(row_idx, 8, QTableWidgetItem(row_data['pmt_serial']))
            self.data_table.setItem(row_idx, 9, QTableWidgetItem(row_data['pmt_batch']))
            
            # Test info
            self.data_table.setItem(row_idx, 10, QTableWidgetItem(row_data['test_name']))
            self.data_table.setItem(row_idx, 11, QTableWidgetItem(row_data['test_fixture']))
            
            # Test Date
            test_date = row_data['test_date']
            date_str = test_date.strftime('%Y-%m-%d %H:%M') if test_date else 'N/A'
            self.data_table.setItem(row_idx, 12, QTableWidgetItem(date_str))
        
        # Resize columns to content
        self.data_table.resizeColumnsToContents()
    
    def show_export_settings(self):
        """Show a dialog with additional export settings."""
        dialog = QDialog(self.main_window)
        dialog.setWindowTitle("Export Settings")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        
        # Font settings
        font_group = QGroupBox("Font Settings")
        font_layout = QFormLayout(font_group)
        
        header_size_spin = QSpinBox()
        header_size_spin.setRange(8, 24)
        header_size_spin.setValue(self.export_style.header_font_size)
        font_layout.addRow("Header Font Size:", header_size_spin)
        
        data_size_spin = QSpinBox()
        data_size_spin.setRange(8, 24)
        data_size_spin.setValue(self.export_style.data_font_size)
        font_layout.addRow("Data Font Size:", data_size_spin)
        
        layout.addWidget(font_group)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.export_style.header_font_size = header_size_spin.value()
            self.export_style.data_font_size = data_size_spin.value()
    
    def _update_export_style(self):
        """Update export style from UI controls."""
        self.export_style.transpose = self.transpose_checkbox.isChecked()
        self.export_style.include_limits = self.include_limits_checkbox.isChecked()
        self.export_style.include_units = self.include_units_checkbox.isChecked()
        self.export_style.include_timestamps = self.include_timestamps_checkbox.isChecked()
        self.export_style.include_plots = self.include_plots_checkbox.isChecked()
        self.export_style.group_header_enabled = self.group_headers_checkbox.isChecked()
        self.export_style.zebra_enabled = self.zebra_checkbox.isChecked()
        self.export_style.conditional_enabled = self.cond_format_checkbox.isChecked()
    
    def export_to_excel(self):
        """Export current data to Excel file."""
        if not self.current_data:
            QMessageBox.warning(
                self.main_window,
                "No Data",
                "Please generate a report first before exporting."
            )
            return
        
        # Get save location
        file_path, _ = QFileDialog.getSaveFileName(
            self.main_window,
            "Export to Excel",
            f"pcba_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Update style from UI
            self._update_export_style()
            
            # Show progress
            progress = QProgressDialog(
                "Exporting to Excel...", "Cancel", 0, 100, self.main_window
            )
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.show()
            QApplication.processEvents()
            
            # Create Excel file
            self._create_excel_file(file_path, progress)
            
            progress.setValue(100)
            progress.close()
            
            QMessageBox.information(
                self.main_window,
                "Export Complete",
                f"Report exported successfully to:\n{file_path}"
            )
            
            logger.info(f"Exported report to {file_path}")
            
        except Exception as e:
            logger.exception("Error exporting to Excel")
            QMessageBox.critical(
                self.main_window,
                "Export Error",
                f"Failed to export: {str(e)}"
            )
    
    def _create_excel_file(self, file_path: str, progress: QProgressDialog):
        """Create the Excel file with styling."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        style = self.export_style
        
        # Helper to convert hex color to openpyxl format
        def hex_to_argb(hex_color: str) -> str:
            return "FF" + hex_color.lstrip('#').upper()
        
        # Create styles
        header_fill = PatternFill(start_color=hex_to_argb(style.header_bg_color),
                                   end_color=hex_to_argb(style.header_bg_color),
                                   fill_type='solid')
        header_font = Font(bold=style.header_font_bold, 
                          color=hex_to_argb(style.header_text_color),
                          size=style.header_font_size)
        
        data_fill = PatternFill(start_color=hex_to_argb(style.data_bg_color),
                                end_color=hex_to_argb(style.data_bg_color),
                                fill_type='solid')
        zebra_fill = PatternFill(start_color=hex_to_argb(style.zebra_color),
                                 end_color=hex_to_argb(style.zebra_color),
                                 fill_type='solid')
        data_font = Font(color=hex_to_argb(style.data_text_color),
                        size=style.data_font_size)
        
        pass_fill = PatternFill(start_color=hex_to_argb(style.pass_color),
                                end_color=hex_to_argb(style.pass_color),
                                fill_type='solid')
        fail_fill = PatternFill(start_color=hex_to_argb(style.fail_color),
                                end_color=hex_to_argb(style.fail_color),
                                fill_type='solid')
        
        border_side = Side(style=style.border_style, color=hex_to_argb(style.border_color))
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        center_align = Alignment(horizontal='center', vertical='center')
        
        progress.setValue(20)
        QApplication.processEvents()
        
        # Prepare data
        if style.transpose:
            self._write_transposed_data(ws, header_fill, header_font, data_fill, zebra_fill,
                                        data_font, pass_fill, fail_fill, border, center_align, style)
        else:
            self._write_normal_data(ws, header_fill, header_font, data_fill, zebra_fill,
                                    data_font, pass_fill, fail_fill, border, center_align, style)
        
        progress.setValue(60)
        QApplication.processEvents()
        
        # Add plots sheet if enabled and plots exist
        if style.include_plots:
            self._add_plots_to_workbook(wb, progress)
        
        progress.setValue(90)
        QApplication.processEvents()
        
        # Save workbook
        wb.save(file_path)
    
    def _write_normal_data(self, ws, header_fill, header_font, data_fill, zebra_fill,
                           data_font, pass_fill, fail_fill, border, center_align, style):
        """Write data in normal orientation (specs as rows, tests as columns)."""
        from openpyxl.utils import get_column_letter
        
        # Build headers
        headers = ["Spec Name"]
        if style.include_units:
            headers.append("Unit")
        if style.include_limits:
            headers.extend(["Lower Limit", "Upper Limit"])
        
        headers.extend(["Measurement", "Status", "PIA Serial", "PIA Part",
                       "PMT Serial", "PMT Batch", "Test Name", "Test Fixture"])
        
        if style.include_timestamps:
            headers.append("Test Date")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Write data rows
        for row_idx, row_data in enumerate(self.current_data, 2):
            col = 1
            
            # Spec Name
            ws.cell(row=row_idx, column=col, value=row_data['spec_name'])
            col += 1
            
            # Unit
            if style.include_units:
                ws.cell(row=row_idx, column=col, value=row_data['unit'])
                col += 1
            
            # Limits
            if style.include_limits:
                lower = row_data['lower_limit']
                ws.cell(row=row_idx, column=col, value=lower if lower is not None else "N/A")
                col += 1
                
                upper = row_data['upper_limit']
                ws.cell(row=row_idx, column=col, value=upper if upper is not None else "N/A")
                col += 1
            
            # Measurement
            meas_col = col
            meas = row_data['measurement']
            ws.cell(row=row_idx, column=col, value=meas if meas is not None else "N/A")
            col += 1
            
            # Status
            status_col = col
            if row_data['passed'] is True:
                ws.cell(row=row_idx, column=col, value="PASS")
            elif row_data['passed'] is False:
                ws.cell(row=row_idx, column=col, value="FAIL")
            else:
                ws.cell(row=row_idx, column=col, value="N/A")
            col += 1
            
            # Device info
            ws.cell(row=row_idx, column=col, value=row_data['pia_serial'])
            col += 1
            ws.cell(row=row_idx, column=col, value=row_data['pia_part'])
            col += 1
            ws.cell(row=row_idx, column=col, value=row_data['pmt_serial'])
            col += 1
            ws.cell(row=row_idx, column=col, value=row_data['pmt_batch'])
            col += 1
            
            # Test info
            ws.cell(row=row_idx, column=col, value=row_data['test_name'])
            col += 1
            ws.cell(row=row_idx, column=col, value=row_data['test_fixture'])
            col += 1
            
            # Test Date
            if style.include_timestamps:
                test_date = row_data['test_date']
                ws.cell(row=row_idx, column=col, value=test_date.strftime('%Y-%m-%d %H:%M') if test_date else "N/A")
                col += 1
            
            # Apply row styling
            row_fill = zebra_fill if (style.zebra_enabled and row_idx % 2 == 0) else data_fill
            
            for c in range(1, col):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = row_fill
                cell.font = data_font
                cell.border = border
                cell.alignment = center_align
            
            # Apply conditional formatting to measurement and status cells
            if style.conditional_enabled:
                if row_data['passed'] is True:
                    ws.cell(row=row_idx, column=meas_col).fill = pass_fill
                    ws.cell(row=row_idx, column=status_col).fill = pass_fill
                elif row_data['passed'] is False:
                    ws.cell(row=row_idx, column=meas_col).fill = fail_fill
                    ws.cell(row=row_idx, column=status_col).fill = fail_fill
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _write_transposed_data(self, ws, header_fill, header_font, data_fill, zebra_fill,
                               data_font, pass_fill, fail_fill, border, center_align, style):
        """Write data in transposed orientation (specs as columns)."""
        from openpyxl.utils import get_column_letter
        
        # Group data by device
        device_data: Dict[str, List[Dict]] = {}
        for row_data in self.current_data:
            device_key = f"{row_data['pia_serial']}_{row_data['pmt_serial']}"
            if device_key not in device_data:
                device_data[device_key] = []
            device_data[device_key].append(row_data)
        
        # Get unique spec names
        spec_names = sorted(set(row['spec_name'] for row in self.current_data))
        
        # Write headers
        headers = ["Device", "Test Date"]
        headers.extend(spec_names)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Write data rows
        row_idx = 2
        for device_key, measurements in device_data.items():
            # Group by test date
            test_groups: Dict[str, Dict[str, Any]] = {}
            for m in measurements:
                date_key = m['test_date'].strftime('%Y-%m-%d %H:%M') if m['test_date'] else 'N/A'
                if date_key not in test_groups:
                    test_groups[date_key] = {'device': device_key}
                test_groups[date_key][m['spec_name']] = m
            
            for date_key, test_data in test_groups.items():
                col = 1
                
                # Device
                ws.cell(row=row_idx, column=col, value=device_key)
                col += 1
                
                # Test Date
                ws.cell(row=row_idx, column=col, value=date_key)
                col += 1
                
                # Spec values
                for spec_name in spec_names:
                    spec_data = test_data.get(spec_name)
                    if spec_data:
                        meas = spec_data['measurement']
                        cell = ws.cell(row=row_idx, column=col, 
                                      value=meas if meas is not None else "N/A")
                        
                        # Conditional formatting
                        if style.conditional_enabled:
                            if spec_data['passed'] is True:
                                cell.fill = pass_fill
                            elif spec_data['passed'] is False:
                                cell.fill = fail_fill
                    else:
                        ws.cell(row=row_idx, column=col, value="N/A")
                    col += 1
                
                # Apply row styling
                row_fill = zebra_fill if (style.zebra_enabled and row_idx % 2 == 0) else data_fill
                for c in range(1, col):
                    cell = ws.cell(row=row_idx, column=c)
                    if cell.fill == PatternFill():  # Only if not already colored by conditional
                        cell.fill = row_fill
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = center_align
                
                row_idx += 1
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _add_plots_to_workbook(self, wb, progress):
        """Add plot images to a separate sheet in the workbook."""
        try:
            from openpyxl.drawing.image import Image as XLImage
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import json
            import tempfile
            import os
        except ImportError as e:
            logger.warning(f"Cannot add plots: {e}")
            return
        
        # Find rows with plot data
        plot_rows = [row for row in self.current_data if row.get('has_plot') and row.get('plot_data')]
        
        if not plot_rows:
            return
        
        # Create plots sheet
        ws_plots = wb.create_sheet("Plots")
        
        row_offset = 1
        for idx, row_data in enumerate(plot_rows):
            try:
                plot_data = json.loads(row_data['plot_data']) if isinstance(row_data['plot_data'], str) else row_data['plot_data']
                
                if not plot_data:
                    continue
                
                # Create plot
                fig, ax = plt.subplots(figsize=(8, 4))
                
                # Handle single or multiple series
                if isinstance(plot_data, list):
                    for series in plot_data:
                        x = series.get('x', [])
                        y = series.get('y', [])
                        label = series.get('label', '')
                        ax.plot(x, y, label=label)
                    ax.legend()
                else:
                    x = plot_data.get('x', [])
                    y = plot_data.get('y', [])
                    ax.plot(x, y)
                
                ax.set_title(f"{row_data['spec_name']} - {row_data['pia_serial']}")
                ax.set_xlabel("X")
                ax.set_ylabel(row_data['unit'] or "Value")
                ax.grid(True, alpha=0.3)
                
                # Save to temp file
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    fig.savefig(tmp.name, dpi=100, bbox_inches='tight', 
                               facecolor='#1e293b', edgecolor='none')
                    plt.close(fig)
                    
                    # Add to worksheet
                    img = XLImage(tmp.name)
                    img.width = 600
                    img.height = 300
                    ws_plots.add_image(img, f'A{row_offset}')
                    
                    # Add label
                    ws_plots.cell(row=row_offset, column=10, 
                                 value=f"{row_data['spec_name']} | {row_data['pia_serial']} | {row_data['test_date']}")
                    
                    row_offset += 20  # Space for next plot
                    
                    # Clean up temp file
                    os.unlink(tmp.name)
                    
            except Exception as e:
                logger.warning(f"Failed to add plot for {row_data['spec_name']}: {e}")
                continue
        
        progress.setValue(80)
        QApplication.processEvents()
    
    def cleanup(self):
        """Clean up resources when page is destroyed."""
        logger.info("ReportsPage cleaned up")
